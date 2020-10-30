#!/usr/bin/env python
# -*- coding:utf-8 -*-
import os
import time
import openpyxl
import jsonpath
import base64
import requests
from urllib.parse import urlencode
# 代理服务器
proxyHost = "http-dyn.abuyun.com"
proxyPort = "9020"

# 代理隧道验证信息
proxyUser = "HM83Z2M4A24KXGLD"
proxyPass = "0A1DBF55E49881D7"

proxyMeta = "http://%(user)s:%(pass)s@%(host)s:%(port)s" % {
    "host": proxyHost,
    "port": proxyPort,
    "user": proxyUser,
    "pass": proxyPass,
}

proxies = {
    "http": proxyMeta,
    "https": proxyMeta,
}

headers = {
    "Accept": "application/json, text/plain, */*",
    "Referer": "https://www.qimai.cn/rank",
    "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:57.0) Gecko/20100101 Firefox/59.0"
}

# 当天时间参数
date = time.strftime("%Y-%m-%d")

# 自定义加密函数
def object_lh(a):
    e = '00000008d78d46a'
    t = len(e)
    n = len(a)
    a = list(a)
    for s in range(n):
        a[s] = chr(ord(a[s]) ^ ord(e[(s + 10) % t]))
    return ''.join(a)

# 组装回调函数
def request_server(params):
    # 提取查询参数值并排序
    s = "".join(sorted([str(v) for v in params.values()]))
    # Base64 Encode
    s = base64.b64encode(bytes(s, encoding="ascii"))
    # 时间差
    t = str(int((time.time() * 1000 - 1515125653845)))
    # 拼接自定义字符串
    s = "@#".join([s.decode(), "/rank/marketRank/", t, "1"])
    # 自定义加密 & Base64 Encode
    s = base64.b64encode(bytes(object_lh(s), encoding="ascii"))
    # 拼接 URL
    params["analysis"] = s.decode()
    url = "https://api.qimai.cn/rank/marketRank/?{}".format(urlencode(params))
    # 发起请求
    return requests.get(url, headers=headers, proxies=proxies)

    # 安卓
def qimai_android_huawei_yingyong(file_path):
    # market = {'华为': 6}
    category = {'影音娱乐': 5, '实用工具': 6, '社交通讯': 7, '教育': 8, '新闻阅读': 9, '拍摄美化': 10, '出行导航': 11, '旅游住宿': 12, "购物比较": 13 , "商务" : 14,'儿童': 15, '金融理财': 16, '运动健康': 17,'便捷生活': 18, '汽车': 19,'主题个性': 20, "美食": 21}
    # 新建工作表
    wb = openpyxl.Workbook()
    for cat_key in category:
        params = {
            "market": 6,
            "category": category[cat_key],
            "date": date,
        }
        # 根据分类创建子表
        ws = wb.create_sheet(cat_key)
        res = request_server(params)
        print(res.status_code)
        jsonobj = res.json()
        print(jsonobj)
        #安卓数据列表
        row_header = ["应用","排名","排名变化","类别","昨日下载量","评分数量","评分结果","最后更新","公司名称"]
        ws.append(row_header)
        applist = jsonpath.jsonpath(jsonobj, '$.rankInfo[*]')
        for appdict in applist:
            row_content = [appdict['appInfo']['appName'],appdict['rankInfo']['ranking'],appdict['rankInfo']['change'],appdict['genre'],appdict['downloadNum'],appdict['appInfo']['app_comment_count'],appdict['appInfo']['app_comment_score'],appdict['releaseTime'],appdict['company']['name']]
            ws.append(row_content)
        time.sleep(1)
        wb.save(file_path+date+'.xlsx')

def qimai_android_huawei_youxi(file_path):
    # market = {'华为': 6}
    category = {'角色扮演': 23, '休闲益智': 24, '经营策略': 25, '体育竞速': 26, '棋牌桌游': 27, '动作射击': 28}
    # 新建工作表
    wb = openpyxl.Workbook()
    for cat_key in category:
        params = {
            "market": 6,
            "category": category[cat_key],
            "date": date,
        }
        # 根据分类创建子表
        ws = wb.create_sheet(cat_key)
        res = request_server(params)
        print(res.status_code)
        jsonobj = res.json()
        print(jsonobj)
        #安卓数据列表
        row_header = ["应用","排名","排名变化","类别","昨日下载量","评分数量","评分结果","最后更新","公司名称"]
        ws.append(row_header)
        applist = jsonpath.jsonpath(jsonobj, '$.rankInfo[*]')
        for appdict in applist:
            row_content = [appdict['appInfo']['appName'],appdict['rankInfo']['ranking'],appdict['rankInfo']['change'],appdict['genre'],appdict['downloadNum'],appdict['appInfo']['app_comment_count'],appdict['appInfo']['app_comment_score'],appdict['releaseTime'],appdict['company']['name']]
            ws.append(row_content)
        time.sleep(1)
        wb.save(file_path+date+'.xlsx')

def qimai_android_xiaomi(file_path):
    # market = {'小米': 4}
    category = {'软件下载榜': 168, '软件飙升榜': 169, '游戏下载榜': 170, '体育竞速': 26, '棋牌桌游': 27, '动作射击': 28}
    # 新建工作表
    wb = openpyxl.Workbook()
    for cat_key in category:
        params = {
            "market": 4,
            "category": category[cat_key],
            "date": date,
        }
        # 根据分类创建子表
        ws = wb.create_sheet(cat_key)
        res = request_server(params)
        print(res.status_code)
        jsonobj = res.json()
        print(jsonobj)
        #安卓数据列表
        row_header = ["应用","排名","排名变化","类别","昨日下载量","评分数量","评分结果","最后更新","公司名称"]
        ws.append(row_header)
        applist = jsonpath.jsonpath(jsonobj, '$.rankInfo[*]')
        for appdict in applist:
            row_content = [appdict['appInfo']['appName'],appdict['rankInfo']['ranking'],appdict['rankInfo']['change'],appdict['genre'],appdict['downloadNum'],appdict['appInfo']['app_comment_count'],appdict['appInfo']['app_comment_score'],appdict['releaseTime'],appdict['company']['name']]
            ws.append(row_content)
        time.sleep(1)
        wb.save(file_path+date+'.xlsx')

def qimai_android_vivo(file_path):
    # market = {'VIVO': 8}
    category = {'应用榜': 4, '单机榜': 5, '网游榜': 6, '热搜榜': 7}
    # 新建工作表
    wb = openpyxl.Workbook()
    for cat_key in category:
        params = {
            "market": 8,
            "category": category[cat_key],
            "date": date,
        }
        # 根据分类创建子表
        ws = wb.create_sheet(cat_key)
        res = request_server(params)
        print(res.status_code)
        jsonobj = res.json()
        print(jsonobj)
        #安卓数据列表
        row_header = ["应用","排名","排名变化","类别","昨日下载量","评分数量","评分结果","最后更新","公司名称"]
        ws.append(row_header)
        applist = jsonpath.jsonpath(jsonobj, '$.rankInfo[*]')
        for appdict in applist:
            row_content = [appdict['appInfo']['appName'],appdict['rankInfo']['ranking'],appdict['rankInfo']['change'],appdict['genre'],appdict['downloadNum'],appdict['appInfo']['app_comment_count'],appdict['appInfo']['app_comment_score'],appdict['releaseTime'],appdict['company']['name']]
            ws.append(row_content)
        time.sleep(1)
        wb.save(file_path+date+'.xlsx')

def qimai_android_oppo(file_path):
    # market = {'OPPO': 9}
    category = {'软件榜': 5, '游戏榜': 6}
    # 新建工作表
    wb = openpyxl.Workbook()
    for cat_key in category:
        params = {
            "market": 9,
            "category": category[cat_key],
            "date": date,
        }
        # 根据分类创建子表
        ws = wb.create_sheet(cat_key)
        res = request_server(params)
        print(res.status_code)
        jsonobj = res.json()
        print(jsonobj)
        #安卓数据列表
        row_header = ["应用","排名","排名变化","类别","昨日下载量","评分数量","评分结果","最后更新","公司名称"]
        ws.append(row_header)
        applist = jsonpath.jsonpath(jsonobj, '$.rankInfo[*]')
        for appdict in applist:
            row_content = [appdict['appInfo']['appName'],appdict['rankInfo']['ranking'],appdict['rankInfo']['change'],appdict['genre'],appdict['downloadNum'],appdict['appInfo']['app_comment_count'],appdict['appInfo']['app_comment_score'],appdict['releaseTime'],appdict['company']['name']]
            ws.append(row_content)
        time.sleep(1)
        wb.save(file_path+date+'.xlsx')

def qimai_android_meizu_yingyong(file_path):
    # market = {'meizu': 7}
    category = {'热门榜': 4, '飙升榜': 6, '新品榜': 7, '游戏榜': 159}
    # 新建工作表
    wb = openpyxl.Workbook()
    for cat_key in category:
        params = {
            "market": 7,
            "category": category[cat_key],
            "date": date,
        }
        # 根据分类创建子表
        ws = wb.create_sheet(cat_key)
        res = request_server(params)
        print(res.status_code)
        jsonobj = res.json()
        print(jsonobj)
        #安卓数据列表
        row_header = ["应用","排名","排名变化","类别","昨日下载量","评分数量","评分结果","最后更新","公司名称"]
        ws.append(row_header)
        applist = jsonpath.jsonpath(jsonobj, '$.rankInfo[*]')
        for appdict in applist:
            row_content = [appdict['appInfo']['appName'],appdict['rankInfo']['ranking'],appdict['rankInfo']['change'],appdict['genre'],appdict['downloadNum'],appdict['appInfo']['app_comment_count'],appdict['appInfo']['app_comment_score'],appdict['releaseTime'],appdict['company']['name']]
            ws.append(row_content)
        time.sleep(1)
        wb.save(file_path+date+'.xlsx')

def qimai_android_meizu_youxi(file_path):
    # market = {'meizu': 7}
    category = {'畅销榜': 5, '热门榜': 173, '新游榜': 174, '免费榜': 175}
    # 新建工作表
    wb = openpyxl.Workbook()
    for cat_key in category:
        params = {
            "market": 7,
            "category": category[cat_key],
            "date": date,
        }
        # 根据分类创建子表
        ws = wb.create_sheet(cat_key)
        res = request_server(params)
        print(res.status_code)
        jsonobj = res.json()
        print(jsonobj)
        #安卓数据列表
        row_header = ["应用","排名","排名变化","类别","昨日下载量","评分数量","评分结果","最后更新","公司名称"]
        ws.append(row_header)
        applist = jsonpath.jsonpath(jsonobj, '$.rankInfo[*]')
        for appdict in applist:
            row_content = [appdict['appInfo']['appName'],appdict['rankInfo']['ranking'],appdict['rankInfo']['change'],appdict['genre'],appdict['downloadNum'],appdict['appInfo']['app_comment_count'],appdict['appInfo']['app_comment_score'],appdict['releaseTime'],appdict['company']['name']]
            ws.append(row_content)
        time.sleep(1)
        wb.save(file_path+date+'.xlsx')

def qimai_android_yingyongbao(file_path):
    # market = {'应用宝': 3}
    category = {'流行榜': 154, '新品榜': 155, '热销榜': 156}
    # 新建工作表
    wb = openpyxl.Workbook()
    for cat_key in category:
        params = {
            "market": 3,
            "category": category[cat_key],
            "date": date,
        }
        # 根据分类创建子表
        ws = wb.create_sheet(cat_key)
        res = request_server(params)
        print(res.status_code)
        jsonobj = res.json()
        print(jsonobj)
        #安卓数据列表
        row_header = ["应用","排名","排名变化","类别","昨日下载量","评分数量","评分结果","最后更新","公司名称"]
        ws.append(row_header)
        applist = jsonpath.jsonpath(jsonobj, '$.rankInfo[*]')
        for appdict in applist:
            row_content = [appdict['appInfo']['appName'],appdict['rankInfo']['ranking'],appdict['rankInfo']['change'],appdict['genre'],appdict['downloadNum'],appdict['appInfo']['app_comment_count'],appdict['appInfo']['app_comment_score'],appdict['releaseTime'],appdict['company']['name']]
            ws.append(row_content)
        time.sleep(1)
        wb.save(file_path+date+'.xlsx')

def qimai_android_baidu_youxi(file_path):
    # market = {'百度': 2}
    category = {'热搜榜': 198, '飙升榜': 199, '新游榜': 200, '单机榜': 201, '网游榜': 202, '精品榜': 203, '预约榜': 204, '休闲': 205, '跑酷': 206, 'MOBA': 207, '角色扮演': 208, '棋牌': 209, '赛车': 210, '射击': 211, '养成': 212, '塔防': 213, '消除': 214, '解密逃脱': 215}
    # 新建工作表
    wb = openpyxl.Workbook()
    for cat_key in category:
        params = {
            "market": 2,
            "category": category[cat_key],
            "date": date,
        }
        # 根据分类创建子表
        ws = wb.create_sheet(cat_key)
        res = request_server(params)
        print(res.status_code)
        jsonobj = res.json()
        print(jsonobj)
        #安卓数据列表
        row_header = ["应用","排名","排名变化","类别","昨日下载量","评分数量","评分结果","最后更新","公司名称"]
        ws.append(row_header)
        applist = jsonpath.jsonpath(jsonobj, '$.rankInfo[*]')
        for appdict in applist:
            row_content = [appdict['appInfo']['appName'],appdict['rankInfo']['ranking'],appdict['rankInfo']['change'],appdict['genre'],appdict['downloadNum'],appdict['appInfo']['app_comment_count'],appdict['appInfo']['app_comment_score'],appdict['releaseTime'],appdict['company']['name']]
            ws.append(row_content)
        time.sleep(1)
        wb.save(file_path+date+'.xlsx')

def qimai_android_baidu_ruanjian(file_path):
    # market = {'百度': 2}
    category = {'热搜榜': 216, '飙升榜': 217, '新锐榜': 218, '视频': 219, '音乐': 220, '社交': 221, '电子书': 222, '直播': 223, '漫画': 224, '购物': 225, '新闻': 226, '拍照': 227, '短视频': 228, '辅导': 229, '儿童': 230, '理财': 231, '旅游': 232, '出行': 233, '母婴': 234, '健康': 235, '办公': 236, '工具': 237, '手机管理': 238, '求职': 239, '美化': 240}
    # 新建工作表
    wb = openpyxl.Workbook()
    for cat_key in category:
        params = {
            "market": 2,
            "category": category[cat_key],
            "date": date,
        }
        # 根据分类创建子表
        ws = wb.create_sheet(cat_key)
        res = request_server(params)
        print(res.status_code)
        jsonobj = res.json()
        print(jsonobj)
        #安卓数据列表
        row_header = ["应用","排名","排名变化","类别","昨日下载量","评分数量","评分结果","最后更新","公司名称"]
        ws.append(row_header)
        applist = jsonpath.jsonpath(jsonobj, '$.rankInfo[*]')
        for appdict in applist:
            row_content = [appdict['appInfo']['appName'],appdict['rankInfo']['ranking'],appdict['rankInfo']['change'],appdict['genre'],appdict['downloadNum'],appdict['appInfo']['app_comment_count'],appdict['appInfo']['app_comment_score'],appdict['releaseTime'],appdict['company']['name']]
            ws.append(row_content)
        time.sleep(1)
        wb.save(file_path+date+'.xlsx')

def qimai_android_360(file_path):
    # market = {'360': 1}
    category = {'飙升榜': 706, '热搜榜': 707, '新锐榜': 708, '月总榜': 709}
    # 新建工作表
    wb = openpyxl.Workbook()
    for cat_key in category:
        params = {
            "market": 1,
            "category": category[cat_key],
            "date": date,
        }
        # 根据分类创建子表
        ws = wb.create_sheet(cat_key)
        res = request_server(params)
        print(res.status_code)
        jsonobj = res.json()
        print(jsonobj)
        #安卓数据列表
        row_header = ["应用","排名","排名变化","类别","昨日下载量","评分数量","评分结果","最后更新","公司名称"]
        ws.append(row_header)
        applist = jsonpath.jsonpath(jsonobj, '$.rankInfo[*]')
        for appdict in applist:
            row_content = [appdict['appInfo']['appName'],appdict['rankInfo']['ranking'],appdict['rankInfo']['change'],appdict['genre'],appdict['downloadNum'],appdict['appInfo']['app_comment_count'],appdict['appInfo']['app_comment_score'],appdict['releaseTime'],appdict['company']['name']]
            ws.append(row_content)
        time.sleep(1)
        wb.save(file_path+date+'.xlsx')

def qimai_android_wandoujia(file_path):
    # market = {'豌豆荚': 5}
    category = {'软件下载榜': 176, '游戏下载榜': 177, '专业榜': 178, '飙升榜': 179, '新游榜': 180}
    # 新建工作表
    wb = openpyxl.Workbook()
    for cat_key in category:
        params = {
            "market": 5,
            "category": category[cat_key],
            "date": date,
        }
        # 根据分类创建子表
        ws = wb.create_sheet(cat_key)
        res = request_server(params)
        print(res.status_code)
        jsonobj = res.json()
        print(jsonobj)
        #安卓数据列表
        row_header = ["应用","排名","排名变化","类别","昨日下载量","评分数量","评分结果","最后更新","公司名称"]
        ws.append(row_header)
        applist = jsonpath.jsonpath(jsonobj, '$.rankInfo[*]')
        for appdict in applist:
            row_content = [appdict['appInfo']['appName'],appdict['rankInfo']['ranking'],appdict['rankInfo']['change'],appdict['genre'],appdict['downloadNum'],appdict['appInfo']['app_comment_count'],appdict['appInfo']['app_comment_score'],appdict['releaseTime'],appdict['company']['name']]
            ws.append(row_content)
        time.sleep(1)
        wb.save(file_path+date+'.xlsx')

def qimai_ios():
        # IOS 免费榜
        params = {
            "brand": "free",
            "country": "cn",
            "date": date,
            "device": "iphone",
            "genre": "6000",
            "page": 1
        }
        # 提取查询参数值并排序
        s = "".join(sorted([str(v) for v in params.values()]))
        # Base64 Encode
        s = base64.b64encode(bytes(s, encoding="ascii"))
        # 时间差
        t = str(int((time.time() * 1000 - 1515125653845)))
        # 拼接自定义字符串
        s = "@#".join([s.decode(), "/rank/index", t, "1"])
        # 自定义加密 & Base64 Encode
        s = base64.b64encode(bytes(object_lh(s), encoding="ascii"))
        # 拼接 URL
        params["analysis"] = s.decode()
        url = "https://api.qimai.cn/rank/index?{}".format(urlencode(params))
        # 发起请求
        res = requests.get(url, headers=headers, proxies=proxies)
        print(res.status_code)
        jsonobj = res.json()
        print(jsonobj)
        # 安卓数据列表
        # applist = jsonpath.jsonpath(jsonobj, '$.rankInfo[*]')
        # print("应用\t排名\t排名变化\t类别\t昨日下载量\t评分数量\t评分结果\t最后更新\t公司名称\n")
        # for appdict in applist:
        #     print(appdict['appInfo']['appName'], end="\t")
        #     print(appdict['rankInfo']['ranking'], end="\t")
        #     print(appdict['rankInfo']['change'], end="\t")
        #     print(appdict['genre'], end="\t")
        #     print(appdict['downloadNum'], end="\t")
        #     print(appdict['appInfo']['app_comment_count'], end="\t")
        #     print(appdict['appInfo']['app_comment_score'], end="\t")
        #     print(appdict['releaseTime'], end="\t")
        #     print(appdict['company']['name'], end="\t")
        #     print("\n")
        # IOS数据列表
        # applist = jsonpath.jsonpath(jsonobj, '$.list[*]')

if __name__ == '__main__':
    dir_path = "../excel/"+date
    # if not os.path.isdir("../excel/"+date):
    if os.path.isdir("../excel/"+date):
        # os.mkdir(dir_path)
        qimai_android_huawei_yingyong(dir_path+'/华为——中国应用')
        time.sleep(20)
        qimai_android_huawei_youxi(dir_path+'/华为——中国游戏')
        time.sleep(20)
        qimai_android_xiaomi(dir_path + '/小米——中国')
        time.sleep(20)
        qimai_android_vivo(dir_path + '/VIVO——中国')
        time.sleep(20)
        qimai_android_oppo(dir_path + '/OPPO——中国')
        time.sleep(20)
        qimai_android_meizu_yingyong(dir_path+'/魅族——中国应用')
        time.sleep(20)
        qimai_android_meizu_youxi(dir_path+'/魅族——中国游戏')
        time.sleep(20)
        qimai_android_yingyongbao(dir_path + '/应用宝——中国')
        time.sleep(20)
        qimai_android_baidu_youxi(dir_path + '/百度——中国游戏')
        time.sleep(20)
        qimai_android_baidu_ruanjian(dir_path + '/百度——中国软件')
        time.sleep(20)
        qimai_android_360(dir_path + '/360——中国')
        time.sleep(20)
        qimai_android_wandoujia(dir_path + '/豌豆荚——中国')
        qimai_ios()
