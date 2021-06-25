import pymongo
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

MONGO_CONNECTION_STRING = 'mongodb://localhost:27017'
MONGO_DB_NAME = 'huawei'
MONGO_COLLECTION_NAME = 'huawei'
client = pymongo.MongoClient(MONGO_CONNECTION_STRING)
db = client['huawei']
collection = db['huawei']

wb = load_workbook("C:/Users/Administrator/Desktop/test.xlsx")
ws = wb["consumptionsum"]
MAX_ROW = ws.max_row
MAX_COL = ws.max_column


def parse_detail(row):
    if row > 1:
        rank = ws.cell(row, 1).value
        title = ws.cell(row, 2).value
        _id = ws.cell(row, 3).value
        account = ws.cell(row, 4).value
        publishTime = ws.cell(row, 5).value
        url = ws.cell(row, 6).value
        exposeCount = ws.cell(row, 7).value
        clickCount = ws.cell(row, 8).value
        clickThroughRate = ws.cell(row, 9).value
        comment = ws.cell(row, 10).value
        share = ws.cell(row, 11).value
    return {

        'rank': int(rank),

        'title': title,

        '_id': _id,

        'account': account,

        'publishTime': publishTime,

        'url': url,

        'exposeCount': int(exposeCount),

        'clickCount': int(clickCount),

        'clickThroughRate': clickThroughRate,

        'comment': int(comment),

        'share': int(share)

    }


def save_data(data):
    collection.update_one({
        '_id': data.get('_id')
    }, {
        '$set': data
    }, upsert=True)


def main():
    i = 2
    while i <= 10:
        data = parse_detail(i)
        save_data(data)
        print(data)
        i += 1


if __name__ == '__main__':
    main()
